from src.core.object_types import format_reporting
from src.core.abstract_report import abstract_report
from src.core.custom_raise import CustomRaise

import openpyxl
import os

class xlsx_report(abstract_report):
    def __init__(self) -> None:
        super().__init__()
        self.__format = format_reporting.XLSX
    
    def create(self, data: list):
        CustomRaise.type_exception("data", data, list)
        if len(data) == 0:
            CustomRaise.operation_exception("Набор данных пуст")

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

        first_model = data[0]
        
        if isinstance(first_model, tuple):
            fields = [f"column_{i}" for i in range(len(first_model))]
        else:
            fields = list(filter(lambda x: not x.startswith("_") and not callable(getattr(first_model, x)), dir(first_model)))
        
        for col_num, field in enumerate(fields, start=1):
            self.sheet.cell(row=1, column=col_num, value=field)
        
        for row_num, row in enumerate(data, start=2):
            for col_num, field in enumerate(fields, start=1):
                value = getattr(row, field)
                if isinstance(value, list):
                    value = ', '.join([str(v) for v in value])
                self.sheet.cell(row=row_num, column=col_num, value=str(value))

    def save(self, directory: str, filename: str):
        full_path = os.path.abspath(os.path.join(os.path.dirname(__file__), directory))

        if not os.path.exists(full_path):
            os.makedirs(full_path)
        
        try:
            file_path = os.path.join(full_path, filename + ".xlsx")
            self.workbook.save(file_path)
            return True
        except:
            return False