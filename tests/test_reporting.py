from src.settings_manager import settings_manager
from src.storage_reposity import storage_reposity
from src.start_service import start_service
from src.core.object_types import format_reporting
from src.reports.csv_report import csv_report
from src.reports.md_report import md_report
from src.reports.report_factory import report_factory
from src.parsers_manager import parsers_manager

import unittest
import os
import json
from docx import Document
import openpyxl

class test_reporting(unittest.TestCase):
    def test_report_create(self):
        
        
        manager = settings_manager()
        manager.open("settings1.json", "../")
        reposity = storage_reposity()
        start = start_service(reposity, manager)
        start.create()
        report = csv_report()
        report.create(reposity.data[ storage_reposity.ranges_key() ])

        # Проверка
        assert report.result != ""  

    def test_factory_create(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings1.json", "../")
        report = report_factory().create( manager )

        # Проверка
        assert report is not None
        assert isinstance(report,  md_report)
    
    def test_report_save_csv(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings (default).json", "../")
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.csv")
        assert os.path.exists(file_path), "Файл отчёта CSV не был создан."
        with open(file_path, 'r', encoding='utf-8') as file:
            assert file.readline() == "id;ingredients;instructions;name;servings;time;\n", "Содержимое заголовка отличается от ожидаемого."
            assert file.readline()[32:80] == ";[Ingredient(name: Пшеничная мука, unit_value: 1", "Содержимое первой строки данных отличается от ожидаемого."
        
    def test_report_save_md(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings1.json", "../")
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.md")
        assert os.path.exists(file_path), "Файл отчёта MD не был создан."
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            assert lines[17] == "**name**: ВАФЛИ ХРУСТЯЩИЕ В ВАФЕЛЬНИЦЕ\n", "Название отличается от ожидаемого."
            assert lines[18] == "**servings**: `10 порций`\n", "Количество порций отличается от ожидаемого."
            assert lines[19] == "**time**: `20 мин`\n", "Время приготовления отличается от ожидаемого."
            assert "- Ingredient(name: Пшеничная мука, unit_value: 100, range: Range(unit: гр, factor: 1), nomenclature: Nomenclature(full_name: Пшеничная мука, group: Group_nomenclature(name: Сырьё), range: Range(unit: гр, factor: 1)))\n" in lines[2], "Ингредиенты отсутствуют или отличаются от ожидаемого."
            assert "- 1. Как испечь вафли хрустящие в вафельнице? Подготовьте необходимые продукты. Из данного количества у меня получилось 8 штук диаметром около 10 см.\n" in lines[8], "Инструкции отсутствуют или отличаются от ожидаемого."

    def test_report_save_json(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings2.json", "../")
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.json")
        assert os.path.exists(file_path), "Файл отчета JSON не был создан."
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            assert data[0]['_recipe__name'] == 'ВАФЛИ ХРУСТЯЩИЕ В ВАФЕЛЬНИЦЕ', "Название отличается от ожидаемого."
            assert data[0]['_recipe__servings'] == '`10 порций`', "Количество порций отличается от ожидаемого."
            assert data[0]['_recipe__time'] == '`20 мин`', "Время приготовления отличается от ожидаемого."
            assert data[0]['_recipe__ingredients'][0]['name'] == "Пшеничная мука", "Ингредиенты отсутствуют или отличаются от ожидаемого."
            assert data[0]['_recipe__instructions'][0] == "1. Как испечь вафли хрустящие в вафельнице? Подготовьте необходимые продукты. Из данного количества у меня получилось 8 штук диаметром около 10 см.", "Инструкции отсутствуют или отличаются от ожидаемого."

    def test_report_save_xml(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings3.json", "../")
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.xml")
        assert os.path.exists(file_path), "Файл отчета XML не был создан."
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            assert "<name>ВАФЛИ ХРУСТЯЩИЕ В ВАФЕЛЬНИЦЕ</name>" in lines[21], "Название отличается от ожидаемого."
            assert "<servings>`10 порций`</servings>" in lines[22], "Количество порций отличается от ожидаемого."
            assert "<time>`20 мин`</time>" in lines[23], "Время приготовления отличается от ожидаемого."
            assert "<element>Ingredient(name: Пшеничная мука, unit_value: 100, range: Range(unit: гр, factor: 1), nomenclature: Nomenclature(full_name: Пшеничная мука, group: Group_nomenclature(name: Сырьё), range: Range(unit: гр, factor: 1)))</element>" in lines[4], "Ингредиенты отсутствуют или отличаются от ожидаемого."
            assert "<element>1. Как испечь вафли хрустящие в вафельнице? Подготовьте необходимые продукты. Из данного количества у меня получилось 8 штук диаметром около 10 см.</element>" in lines[11], "Инструкции отсутствуют или отличаются от ожидаемого."

    def test_report_save_docx(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings3.json", "../")
        manager.settings.report_format = format_reporting.DOCX
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.docx")
        assert os.path.exists(file_path), "Файл отчета DOCX не был создан."
        document = Document(file_path)
        paragraphs = [p.text for p in document.paragraphs]
        assert "Name: ВАФЛИ ХРУСТЯЩИЕ В ВАФЕЛЬНИЦЕ" in paragraphs, "Название отличается от ожидаемого."
        assert "Servings: `10 порций`" in paragraphs, "Количество порций отличается от ожидаемого."
        assert "Time: `20 мин`" in paragraphs, "Время приготовления отличается от ожидаемого."
        
    def test_report_save_xlsx(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings3.json", "../")
        manager.settings.report_format = format_reporting.XLSX
        reposity = storage_reposity()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")

        # Проверка
        file_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/reports")), "receipts_report.xlsx")
        assert os.path.exists(file_path), "Файл отчета XLSX не был создан."
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        assert sheet.cell(row=2, column=4).value == "ВАФЛИ ХРУСТЯЩИЕ В ВАФЕЛЬНИЦЕ", "Название отличается от ожидаемого."
        assert sheet.cell(row=2, column=5).value == "`10 порций`", "Количество порций отличается от ожидаемого."
        assert sheet.cell(row=2, column=6).value == "`20 мин`", "Время приготовления отличается от ожидаемого."

    def test_custom_report_mapping(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings (default).json", "../")
        manager.settings.report_mapping = {format_reporting.CSV: md_report}
        report = report_factory().create(manager)
        
        # Проверка
        assert report is not None
        assert isinstance(report, md_report), "Фабрика отчётов не использовала кастомный маппинг."

    def test_deserialize_json(self):
        # Подготовка
        manager = settings_manager()
        manager.open("settings2.json", "../")
        reposity = storage_reposity()
        start = start_service(reposity, manager)
        start.create()
        report = report_factory().create(manager)
        report.create(reposity.data[storage_reposity.receipts_key()])
        report.save("../../data/reports", "receipts_report")
        data = parsers_manager.parse_json("receipts_report.json")

        # Проверка
        assert start.data["receipts"][0].name == data[0].name, f"{data[0]}"
        assert start.data["receipts"][0].servings == data[0].servings
        assert start.data["receipts"][0].ingredients[0].name == data[0].ingredients[0].name
        assert start.data["receipts"][0].ingredients[0].unit_value == data[0].ingredients[0].unit_value
        assert start.data["receipts"][0].ingredients[0].range.unit_range == data[0].ingredients[0].range.unit_range
        assert start.data["receipts"][0].ingredients[0].nomenclature.group.name == data[0].ingredients[0].nomenclature.group.name
        assert start.data["receipts"][0].time == data[0].time
        assert start.data["receipts"][0].instructions[0] == data[0].instructions[0]